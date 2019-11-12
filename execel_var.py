import openpyxl

ex_name = 'db_check.xlsx'  # excel file name (database)
ex = openpyxl.load_workbook(ex_name)  # opens the excel

sheet = ex["Sheet1"]  # sheet one for main database


# b_sheet = ex['sheet2']  # sheet two for skipped people

# b_sheet.cell(row=b_sheet.max_row, column=1).value = 'Not on the list'
def create_sheet():
    if 'sheet2' not in ex.sheetnames:
        ex.create_sheet('sheet2')
        ex.save(ex_name)
    b_sheet = ex['sheet2']
    return b_sheet

