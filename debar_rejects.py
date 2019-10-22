import openpyxl
import os
from collections import defaultdict

ex_name = 'db_check.xlsx'
ex = openpyxl.load_workbook(ex_name)
sheet = ex["Sheet1"]
b_sheet = ex['sheet2']
dd = defaultdict(list)
counter = 0


# finds the empty row cell
def check():

    b_sheet.cell(row=b_sheet.max_row+1, column=1).value = 'Hiiii'


# value = b_sheet.cell(row=b_sheet.max_row +1, column=1).value)

    ex.save(ex_name)

check()


def check_row_isempty():
    # for num, r in enumerate(b_sheet.iter_rows()):
    for row_cells in b_sheet.iter_rows(min_col=1, max_col=1):
        for cell in row_cells:
            if cell.value is None:
                return row_cells
            else:
                break

    print(row_cells)
